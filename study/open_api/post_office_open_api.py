# import requests
# import re
# import openpyxl
# from urllib.parse import quote
#
# # 엑셀 파일 경로
# input_file = "C:/workSpace/pythonBasic/data/zip_code_open_api.xlsx"
#
# # 요청할 URL
# url = 'http://openapi.epost.go.kr/postal/retrieveNewAdressAreaCdSearchAllService/retrieveNewAdressAreaCdSearchAllService/getNewAddressListAreaCdSearchAll'
#
# # 서비스 키
# service_key = 'SJpTwB+9NLadh2zz0JcxICsLgq3BYrPpGHh9M9cLDuWS4ifD3Z6/K4eyHmegOkiMGXVYB5TpJ0QKkoThu410hQ=='  # 여기에 실제 서비스 키를 넣으세요
#
# # 엑셀 파일 열기
# wb = openpyxl.load_workbook(input_file)
# sheet = wb.active
#
# # 결과를 저장할 엑셀 파일 생성
# output_file = "C:/workSpace/pythonBasic/data/zip_code_results.xlsx"
# output_wb = openpyxl.Workbook()
# output_sheet = output_wb.active
# output_sheet.append(['Address', 'ZipCode', 'Response'])
#
# # 주소에서 '로', '길', '동'으로 끝나는 부분을 추출하는 함수
# def extract_address_part(address):
#     # '로', '길', '동'으로 끝나는 패턴을 찾는 정규 표현식
#     pattern = r'([가-힣a-zA-Z0-9]+(?:로|길|동))'
#
#     # 정규 표현식으로 패턴에 맞는 부분을 찾기
#     match = re.search(pattern, address)
#
#     if match:
#         # 정규 표현식이 매칭된 부분을 반환
#         print(match.group(0))
#         return match.group(0)
#     return None
#
# # 엑셀에서 주소를 읽고 API 요청 보내기
# for row in sheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더라고 가정하고 두 번째 행부터 시작
#     address = row[0]  # 주소가 첫 번째 열에 있다고 가정
#
#     if isinstance(address, str):
#         # 주소에서 '로', '길', '동'으로 끝나는 부분 추출
#         srchwrd = extract_address_part(address)
#
#         if srchwrd:
#             # 요청 파라미터 설정
#             params = {
#                 'serviceKey': service_key,
#                 'srchwrd': srchwrd,  # 주소 인코딩
#                 'countPerPage': '50',
#                 'currentPage': '1'
#             }
#
#             # API 요청 보내기
#             response = requests.get(url, params=params)
#
#             if response.status_code == 200:
#                 # 응답에서 필요한 정보를 추출 (예: 우편번호)
#                 response_data = response.content.decode('utf-8')
#                 print(f"응답 데이터: {response_data}")
#
#                 # 여기에 적절한 파싱 후 응답 내용을 추출하여 저장할 수 있음
#                 zip_code = "ExtractedZipCode"  # 실제로 응답에서 우편번호를 추출하는 코드 필요
#                 output_sheet.append([address, zip_code, response_data])
#             else:
#                 print(f"요청 실패. 상태 코드: {response.status_code}")
#
# # 결과를 엑셀 파일로 저장
# output_wb.save(output_file)
# print(f"결과가 {output_file}에 저장되었습니다.")

import requests
import re
import openpyxl
import xml.etree.ElementTree as ET

# 엑셀 파일 경로
input_file = "C:/workSpace/pythonBasic/data/zip_code_open_api.xlsx"

# 요청할 URL
url = 'http://openapi.epost.go.kr/postal/retrieveNewAdressAreaCdSearchAllService/retrieveNewAdressAreaCdSearchAllService/getNewAddressListAreaCdSearchAll'

# 서비스 키
service_key = 'SJpTwB+9NLadh2zz0JcxICsLgq3BYrPpGHh9M9cLDuWS4ifD3Z6/K4eyHmegOkiMGXVYB5TpJ0QKkoThu410hQ=='  # 여기에 실제 서비스 키를 넣으세요

# 엑셀 파일 열기
wb = openpyxl.load_workbook(input_file)
sheet = wb.active

# 결과를 저장할 엑셀 파일 생성
output_file = "C:/workSpace/pythonBasic/data/zip_code_results.xlsx"
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
output_sheet.append(['Address', 'ZipCode'])  # Header only for address and zip code

# 주소에서 '로', '길', '동'으로 끝나는 부분을 추출하는 함수
def extract_address_part(address):
    # '로', '길', '동'으로 끝나는 패턴을 찾는 정규 표현식
    pattern = r'([가-힣a-zA-Z0-9]+(?:로|길|동))'

    # 정규 표현식으로 패턴에 맞는 부분을 찾기
    match = re.search(pattern, address)

    if match:
        # 정규 표현식이 매칭된 부분을 반환
        return match.group(0)
    return None

# 엑셀에서 주소를 읽고 API 요청 보내기
for row in sheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더라고 가정하고 두 번째 행부터 시작
    address = row[0]  # 주소가 첫 번째 열에 있다고 가정

    if isinstance(address, str):
        # 주소에서 '로', '길', '동'으로 끝나는 부분 추출
        srchwrd = extract_address_part(address)

        if srchwrd:
            # 요청 파라미터 설정
            params = {
                'serviceKey': service_key,
                'srchwrd': srchwrd,  # 주소 인코딩
                'countPerPage': '50',
                'currentPage': '1'
            }

            # API 요청 보내기
            response = requests.get(url, params=params)

            if response.status_code == 200:
                # 응답에서 XML 데이터를 파싱
                tree = ET.ElementTree(ET.fromstring(response.content))
                root = tree.getroot()

                # totalCount 값을 추출 (없을 수 있으므로 None 체크)
                total_count_element = root.find(".//totalCount")
                if total_count_element is not None and total_count_element.text:
                    try:
                        total_count = int(total_count_element.text)
                        total_page = (total_count // 50) + (1 if total_count % 50 > 0 else 0)

                        # 페이지별로 데이터를 가져옴
                        for page in range(1, total_page + 1):
                            params['currentPage'] = str(page)
                            response = requests.get(url, params=params)

                            if response.status_code == 200:
                                tree = ET.ElementTree(ET.fromstring(response.content))
                                root = tree.getroot()

                                # XML에서 우편번호와 주소 추출
                                for address_item in root.findall(".//newAddressListAreaCdSearchAll"):
                                    zip_code = address_item.find("zipNo").text  # 우편번호
                                    lnm_address = address_item.find("lnmAdres").text  # 법정동 주소
                                    rn_address = address_item.find("rnAdres").text  # 도로명 주소

                                    # 주소가 있으면 법정동 주소 또는 도로명 주소를 저장
                                    if lnm_address:
                                        output_sheet.append([lnm_address, zip_code])
                                    elif rn_address:
                                        output_sheet.append([rn_address, zip_code])
                            else:
                                print(f"페이지 {page} 요청 실패. 상태 코드: {response.status_code}")
                    except ValueError:
                        print(f"totalCount 값이 유효하지 않습니다: {total_count_element.text}")
                else:
                    print("totalCount가 없거나 비어 있습니다.")
            else:
                print(f"요청 실패. 상태 코드: {response.status_code}")

# 결과를 엑셀 파일로 저장
output_wb.save(output_file)
print(f"결과가 {output_file}에 저장되었습니다.")
